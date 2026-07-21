"""
reports/excel_export.py
------------------------
Excel export logic for the Gravity Adjustment Software (Phase 8).

This module contains NO GUI code -- it is pure export logic, called
from gui.py's export slot. It takes the same pandas DataFrames already
displayed in the app's tables and writes them to a formatted .xlsx
workbook using openpyxl.

This is a REPORT export (a snapshot of already-computed values), not a
live editable financial model -- so cells are written as plain values,
not formulas. If a future requirement needs the workbook to recalculate
when edited (e.g. a surveyor tweaking a raw reading and expecting
Drift/GValue to update), that would be a deliberate follow-up change,
not this module's current scope.

Sheets produced (only for data that's actually available -- a sheet is
skipped entirely if its source data is None):
    "Observations"            <- raw imported data (Station, Time, Reading)
    "Drift Corrected Results" <- DriftCorrector.compute() output, plus a
                                  small summary block (Known G Value,
                                  Total Time, Total Drift, Drift Rate)
                                  pulled from the DataFrame's .attrs
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.drift import format_minutes_to_clock


class ExcelExportError(Exception):
    """Raised when the Excel export cannot be completed."""
    pass


# ----------------------------------------------------------------------
# Shared styling constants (kept consistent with the app's navy/blue
# professional theme used in gui.py's _apply_professional_style()).
# ----------------------------------------------------------------------
FONT_NAME = "Calibri"
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(name=FONT_NAME, color="2C3E50", bold=True, size=14)
LABEL_FONT = Font(name=FONT_NAME, color="2C3E50", bold=True, size=10)
VALUE_FONT = Font(name=FONT_NAME, color="1A2330", size=10)
BODY_FONT = Font(name=FONT_NAME, color="1A2330", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="B8C0CC"),
    right=Side(style="thin", color="B8C0CC"),
    top=Side(style="thin", color="B8C0CC"),
    bottom=Side(style="thin", color="B8C0CC"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def export_to_excel(
    filepath: str,
    observation_data=None,
    drift_corrected_data=None,
    known_g_value: float = None,
) -> None:
    """
    Write a formatted .xlsx workbook containing whichever of the
    supplied datasets are not None.

    Args:
        filepath: destination path, e.g. "/path/to/report.xlsx".
        observation_data: raw imported DataFrame (Station, Time,
            Reading), or None to skip this sheet.
        drift_corrected_data: DriftCorrector.compute() output
            DataFrame (Station, MeanTime, MeanReading, Drift,
            CorrectedReading, DeltaG, GValue), or None to skip this
            sheet. If it has drift-rate attrs (total_time_minutes,
            total_drift, drift_rate_per_minute), those are written as
            a summary block above the table.
        known_g_value: the manually-entered known G value used for
            this run, included in the summary block if provided.

    Raises:
        ExcelExportError: if neither dataset is provided (nothing to
            export), or if writing the file fails for any reason
            (e.g. invalid path, permissions).
    """
    if observation_data is None and drift_corrected_data is None:
        raise ExcelExportError(
            "Nothing to export -- import a file and/or run drift "
            "correction before exporting to Excel."
        )

    workbook = Workbook()
    # Remove the default blank sheet; we add named sheets explicitly below.
    workbook.remove(workbook.active)

    try:
        if observation_data is not None:
            _write_observations_sheet(workbook, observation_data)

        if drift_corrected_data is not None:
            _write_drift_results_sheet(workbook, drift_corrected_data, known_g_value)

        workbook.save(filepath)
    except ExcelExportError:
        raise
    except Exception as exc:
        raise ExcelExportError(f"Failed to write Excel file: {exc}") from exc


# ------------------------------------------------------------------
# INTERNAL SHEET BUILDERS
# ------------------------------------------------------------------
def _write_observations_sheet(workbook, observation_data):
    """Write the raw imported observations to an "Observations" sheet."""
    sheet = workbook.create_sheet("Observations")
    _write_title(sheet, "Imported Observations", num_columns=len(observation_data.columns))

    header_row = 3
    _write_dataframe(
        sheet,
        observation_data,
        start_row=header_row,
        float_format="{:.4f}",
    )


def _write_drift_results_sheet(workbook, drift_corrected_data, known_g_value):
    """
    Write the drift-corrected results to a "Drift Corrected Results"
    sheet, with a summary block (known G value, total time/drift/rate,
    pulled from the DataFrame's .attrs when present) above the table.
    """
    sheet = workbook.create_sheet("Drift Corrected Results")
    num_columns = len(drift_corrected_data.columns)
    _write_title(sheet, "Drift Corrected Results", num_columns=num_columns)

    summary_row = 3
    summary_row = _write_summary_block(sheet, drift_corrected_data, known_g_value, start_row=summary_row)

    header_row = summary_row + 1
    _write_dataframe(
        sheet,
        drift_corrected_data,
        start_row=header_row,
        float_format="{:.6f}",
        time_column="MeanTime",
    )


def _write_summary_block(sheet, drift_corrected_data, known_g_value, start_row):
    """
    Write a small label/value summary block. Only includes rows for
    values that are actually available (known_g_value, and/or the
    total_time_minutes/total_drift/drift_rate_per_minute attrs set by
    DriftCorrector.compute()). Returns the next free row after the block.
    """
    attrs = drift_corrected_data.attrs
    summary_items = []

    if known_g_value is not None:
        summary_items.append(("Known G Value:", f"{known_g_value:.6f}"))
    if "total_time_minutes" in attrs:
        summary_items.append(("Total Time (min):", f"{attrs['total_time_minutes']:.2f}"))
    if "total_drift" in attrs:
        summary_items.append(("Total Drift:", f"{attrs['total_drift']:.6f}"))
    if "drift_rate_per_minute" in attrs:
        summary_items.append(("Drift Rate (per min):", f"{attrs['drift_rate_per_minute']:.6f}"))

    if not summary_items:
        return start_row

    row = start_row
    for label, value in summary_items:
        label_cell = sheet.cell(row=row, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.alignment = LEFT

        value_cell = sheet.cell(row=row, column=2, value=value)
        value_cell.font = VALUE_FONT
        value_cell.alignment = LEFT
        row += 1

    return row + 1  # one blank row after the summary block


def _write_title(sheet, title_text, num_columns):
    """Write a bold title in row 1, merged across the table's column span."""
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(num_columns, 1))
    title_cell = sheet.cell(row=1, column=1, value=title_text)
    title_cell.font = TITLE_FONT
    title_cell.alignment = LEFT


def _write_dataframe(sheet, dataframe, start_row, float_format="{:.4f}", time_column=None):
    """
    Write a DataFrame's header + rows starting at `start_row`, with
    header styling (navy fill, white bold text), thin borders on every
    cell, center alignment, and column widths sized to fit content.

    Args:
        time_column: if given, that column's values are formatted via
            core.drift.format_minutes_to_clock() (H:MM) instead of the
            raw float, matching the app's own results-table display
            convention (populate_results_table() in gui.py).
    """
    columns = [str(c) for c in dataframe.columns]

    for col_idx, column_name in enumerate(columns, start=1):
        cell = sheet.cell(row=start_row, column=col_idx, value=column_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    column_widths = [len(name) + 2 for name in columns]

    for row_offset, (_, row) in enumerate(dataframe.iterrows(), start=1):
        excel_row = start_row + row_offset
        for col_idx, column_name in enumerate(columns, start=1):
            value = row[column_name]

            if column_name == time_column and isinstance(value, (int, float)):
                text = format_minutes_to_clock(value)
            elif isinstance(value, float):
                text = float_format.format(value)
            else:
                text = str(value)

            cell = sheet.cell(row=excel_row, column=col_idx, value=text)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

            column_widths[col_idx - 1] = max(column_widths[col_idx - 1], len(text) + 2)

    for col_idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(width, 40)